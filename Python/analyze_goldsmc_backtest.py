"""
Analyse complète du backtest GoldSMC v5 depuis fichier Excel MT5.

Usage:
    python Python/analyze_goldsmc_backtest.py "Backtest_report/ReportTester-5775742_XAUUSD_25 mai bis.xlsx"
"""

import sys
import io
from pathlib import Path
from datetime import datetime

# Fix Windows encoding
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ Module openpyxl manquant")
    print("   Installer: pip install openpyxl")
    sys.exit(1)


def parse_value(text, value_type="float"):
    """Parse une valeur numérique depuis texte."""
    if not text:
        return None

    try:
        # Nettoyer le texte
        cleaned = str(text).strip().replace(",", "").replace("$", "").replace("%", "")

        if value_type == "float":
            return float(cleaned)
        elif value_type == "int":
            return int(float(cleaned))

    except (ValueError, AttributeError):
        return None


def extract_metrics_from_excel(excel_path: str):
    """Extrait toutes les métriques du backtest MT5 Excel (format français)."""

    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    metrics = {
        "file": Path(excel_path).name,
        "analyzed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

    # Scanner lignes pour extraire métriques
    for row_idx in range(1, ws.max_row + 1):
        # Récupérer valeurs colonnes 1-8
        col1 = ws.cell(row=row_idx, column=1).value
        col2 = ws.cell(row=row_idx, column=2).value
        col3 = ws.cell(row=row_idx, column=3).value
        col4 = ws.cell(row=row_idx, column=4).value
        col5 = ws.cell(row=row_idx, column=5).value
        col6 = ws.cell(row=row_idx, column=6).value
        col7 = ws.cell(row=row_idx, column=7).value
        col8 = ws.cell(row=row_idx, column=8).value

        # Ne pas skip si col1 vide: certaines lignes ont des données en col5/col8 seulement
        text = str(col1).lower() if col1 else ""
        text5 = str(col5).lower() if col5 else ""

        # Skip ligne si col1 ET col5 sont vides
        if not text and not text5:
            continue

        # === CONFIGURATION (valeurs en col4) ===
        if "symbole:" in text and col4:
            metrics["symbol"] = str(col4).strip()

        if "période:" in text and col4:
            # Format: "H1 (2022.01.01 - 2024.01.01)"
            period_str = str(col4).strip()
            metrics["timeframe"] = period_str.split("(")[0].strip() if "(" in period_str else period_str
            if "(" in period_str:
                dates = period_str.split("(")[1].replace(")", "").strip()
                metrics["test_period"] = dates

        if "dépôt initial:" in text and col4 is not None:
            metrics["initial_deposit"] = parse_value(col4, "float")

        if "levier:" in text and col4:
            metrics["leverage"] = str(col4).strip()

        # === RÉSULTATS (valeur en col4, pas col2) ===
        if "profit total net:" in text and col4 is not None:
            metrics["net_profit"] = parse_value(col4, "float")

        if "profit brut:" in text and col4 is not None:
            metrics["gross_profit"] = parse_value(col4, "float")

        if "perte brut:" in text and col4 is not None:
            metrics["gross_loss"] = parse_value(col4, "float")

        if "facteur de profit:" in text and col4 is not None:
            metrics["profit_factor"] = parse_value(col4, "float")

        if "facteur de récupération:" in text and col4 is not None:
            metrics["recovery_factor"] = parse_value(col4, "float")

        if "ratio de sharpe:" in text and col4 is not None:
            metrics["sharpe_ratio"] = parse_value(col4, "float")

        if "nb trades:" in text and col4 is not None:
            metrics["total_trades"] = parse_value(col4, "int")

        if "opérations au total:" in text and col4 is not None:
            metrics["total_deals"] = parse_value(col4, "int")

        # Drawdown maximal (ligne 73: "Solde Drawdown Maximal:" text5, valeur col8)
        if "solde drawdown maximal:" in text5 and col8:
            dd_val = str(col8).strip()
            if "(" in dd_val and "%" in dd_val:
                # Format: "30.38 (6.08%)"
                metrics["max_dd_abs"] = parse_value(dd_val.split("(")[0], "float")
                metrics["max_dd_pct"] = parse_value(dd_val.split("(")[1].replace("%", "").replace(")", ""), "float")

        # Drawdown relatif (ligne 74 - backup si maximal non trouvé)
        if "solde drawdown relatif:" in text5 and col8 and "max_dd_pct" not in metrics:
            dd_val = str(col8).strip()
            if "%" in dd_val and "(" in dd_val:
                # Format: "6.08% (30.38)"
                metrics["max_dd_pct"] = parse_value(dd_val.split("%")[0], "float")
                if "(" in dd_val:
                    metrics["max_dd_abs"] = parse_value(dd_val.split("(")[1].replace(")", ""), "float")

        # Positions courtes (col5 contient le label, col8 la valeur)
        if "positions courtes" in text5 and "(gagnées %):" in text5 and col8:
            shorts = str(col8).strip()
            if "(" in shorts and "%" in shorts:
                short_count = parse_value(shorts.split("(")[0], "int")
                short_pct = parse_value(shorts.split("(")[1].replace("%", "").replace(")", ""), "float")
                metrics["short_positions"] = short_count
                metrics["short_win_pct"] = short_pct

        # Positions gagnantes (col5 label, col8 valeur)
        if "positions gagnantes" in text5 and "(% du total):" in text5 and col8:
            wins = str(col8).strip()
            if "(" in wins and "%" in wins:
                win_count = parse_value(wins.split("(")[0], "int")
                win_pct = parse_value(wins.split("(")[1].replace("%", "").replace(")", ""), "float")
                metrics["winning_trades"] = win_count
                metrics["win_rate"] = win_pct

        # Plus large position (ligne 86: text5 label, col8 valeur)
        if "plus large position gagnante:" in text5 and col8 is not None:
            metrics["largest_win"] = parse_value(col8, "float")

        # Moyenne position gagnante (ligne 87: text5 label, col8 valeur)
        if "moyenne position gagnante:" in text5 and col8 is not None:
            metrics["avg_win"] = parse_value(col8, "float")

        # Magic Number (col4 contient "MagicNumber=20260524")
        text4 = str(col4).lower() if col4 else ""
        if col4 and "magicnumber=" in text4:
            magic = str(col4).replace("MagicNumber=", "").replace("magicnumber=", "").strip()
            metrics["magic_number"] = parse_value(magic, "int")

    wb.close()

    # Si MagicNumber pas trouvé, chercher dans tous les params
    if "magic_number" not in metrics:
        # Re-open pour chercher MagicNumber spécifiquement
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        for row_idx in range(1, min(100, ws.max_row + 1)):
            for col_idx in range(1, 10):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val and "magicnumber=" in str(cell_val).lower():
                    magic_str = str(cell_val).replace("MagicNumber=", "").replace("magicnumber=", "").strip()
                    magic_num = parse_value(magic_str, "int")
                    if magic_num:
                        metrics["magic_number"] = magic_num
                        break
            if "magic_number" in metrics:
                break
        wb.close()

    return metrics


def calculate_derived_metrics(metrics):
    """Calcule les métriques dérivées."""
    derived = {}

    # ROI
    if "net_profit" in metrics and "initial_deposit" in metrics and metrics["initial_deposit"]:
        roi = (metrics["net_profit"] / metrics["initial_deposit"]) * 100
        derived["roi_pct"] = round(roi, 2)

    # Balance finale
    if "net_profit" in metrics and "initial_deposit" in metrics:
        derived["final_balance"] = metrics["initial_deposit"] + metrics["net_profit"]

    # Perte moyenne
    if "gross_loss" in metrics and "total_trades" in metrics and metrics["total_trades"]:
        # Trades perdants = total - gagnants
        winning = metrics.get("winning_trades", 0)
        losing = metrics["total_trades"] - winning
        if losing > 0:
            derived["avg_loss"] = round(metrics["gross_loss"] / losing, 2)
            derived["losing_trades"] = losing

    return derived


def print_report(metrics, derived):
    """Affiche le rapport complet."""

    print(f"\n{'='*80}")
    print(f"  ANALYSE BACKTEST GOLDSMC")
    print(f"{'='*80}")
    print(f"Fichier: {metrics.get('file', 'N/A')}")
    print(f"Analysé: {metrics.get('analyzed_at', 'N/A')}\n")

    # CONFIGURATION
    print(f"{'─'*80}")
    print(f"  CONFIGURATION")
    print(f"{'─'*80}")
    print(f"Symbole:         {metrics.get('symbol', 'N/A')}")
    print(f"Période:         {metrics.get('timeframe', 'N/A')} | {metrics.get('test_period', 'N/A')}")
    print(f"Dépôt initial:   ${metrics.get('initial_deposit', 0):.2f}")
    print(f"Levier:          {metrics.get('leverage', 'N/A')}")
    print(f"Magic Number:    {metrics.get('magic_number', 'N/A')}\n")

    # RÉSULTATS GLOBAUX
    print(f"{'─'*80}")
    print(f"  RÉSULTATS GLOBAUX")
    print(f"{'─'*80}")
    net = metrics.get('net_profit', 0)
    gross_p = metrics.get('gross_profit', 0)
    gross_l = metrics.get('gross_loss', 0)

    print(f"Profit Net:      ${net:.2f}")
    print(f"Bénéfice Brut:   ${gross_p:.2f}")
    print(f"Perte Brute:     ${gross_l:.2f}")
    print(f"Balance Finale:  ${derived.get('final_balance', 0):.2f}")
    print(f"ROI:             {derived.get('roi_pct', 0):.2f}%\n")

    # MÉTRIQUES CLÉS
    print(f"{'─'*80}")
    print(f"  MÉTRIQUES CLÉS")
    print(f"{'─'*80}")
    pf = metrics.get('profit_factor', 0)
    rf = metrics.get('recovery_factor', 0)
    sr = metrics.get('sharpe_ratio', 0)

    print(f"Profit Factor:    {pf:.3f}")
    print(f"Recovery Factor:  {rf:.3f}")
    sharpe_str = f"{sr:.3f}" if sr else "N/A"
    print(f"Sharpe Ratio:     {sharpe_str}\n")

    # DRAWDOWN
    print(f"{'─'*80}")
    print(f"  DRAWDOWN")
    print(f"{'─'*80}")
    dd_abs = metrics.get('max_dd_abs', 0)
    dd_pct = metrics.get('max_dd_pct', 0)

    print(f"Max Drawdown:     ${dd_abs:.2f} ({dd_pct:.2f}%)\n")

    # TRADES
    print(f"{'─'*80}")
    print(f"  TRADES")
    print(f"{'─'*80}")
    total = metrics.get('total_trades', 0)
    wins = metrics.get('winning_trades', 0)
    win_rate = metrics.get('win_rate', 0)

    print(f"Total Trades:      {total}")
    print(f"Trades Gagnants:   {wins} ({win_rate:.1f}%)")
    print(f"Trades Perdants:   {derived.get('losing_trades', 0)}\n")

    print(f"Plus Gros Gain:    ${metrics.get('largest_win', 0):.2f}")
    print(f"Gain Moyen:        ${metrics.get('avg_win', 0):.2f}")
    print(f"Perte Moyenne:     ${derived.get('avg_loss', 0):.2f}\n")

    # OBJECTIFS GOLDSMC V5
    print(f"{'─'*80}")
    print(f"  COMPARAISON OBJECTIFS GOLDSMC V5")
    print(f"{'─'*80}")

    targets = [
        ("Profit Factor ≥ 2.0", pf, 2.0),
        ("Max Drawdown ≤ 20%", dd_pct, 20.0, True),  # True = inverse (plus bas = mieux)
        ("Recovery Factor ≥ 3.0", rf, 3.0),
        ("Win Rate ≥ 50%", win_rate, 50.0),
    ]

    for name, value, target, *inverse in targets:
        is_inverse = inverse[0] if inverse else False

        if is_inverse:
            passed = value <= target
            emoji = "✅" if passed else "❌"
        else:
            passed = value >= target
            emoji = "✅" if passed else "❌"

        print(f"{emoji} {name:30s}  | Actuel: {value:.2f} | Cible: {target:.2f}")

    print(f"\n{'='*80}\n")


def main():
    if len(sys.argv) < 2:
        print("Usage: python analyze_goldsmc_backtest.py <fichier.xlsx>")
        sys.exit(1)

    excel_file = sys.argv[1]

    if not Path(excel_file).exists():
        print(f"❌ Fichier introuvable: {excel_file}")
        sys.exit(1)

    try:
        print("📊 Extraction des métriques...")
        metrics = extract_metrics_from_excel(excel_file)

        print("🔢 Calcul des métriques dérivées...")
        derived = calculate_derived_metrics(metrics)

        print_report(metrics, derived)

        sys.exit(0)

    except Exception as e:
        print(f"\n❌ Erreur: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
