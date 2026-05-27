"""
Extrait les données d'un backtest MT5 Excel

Usage:
    python Python/extract_backtest_excel.py "Backtest_report/ReportTester-5775742_XAUUSD_25 mai bis.xlsx"
"""

import sys
import io
from pathlib import Path

# Fix Windows encoding
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ Module openpyxl manquant")
    print("   Installer: pip install openpyxl")
    sys.exit(1)


def extract_backtest_summary(excel_path: str):
    """Extrait le résumé du backtest depuis Excel MT5."""

    wb = load_workbook(excel_path, data_only=True)

    print(f"\n{'='*70}")
    print(f"  ANALYSE BACKTEST MT5")
    print(f"{'='*70}")
    print(f"Fichier: {Path(excel_path).name}\n")

    # Lister les feuilles
    print(f"📊 Feuilles disponibles: {wb.sheetnames}\n")

    # Analyser la première feuille
    ws = wb.active
    print(f"Feuille active: {ws.title}")
    print(f"Dimensions: {ws.max_row} lignes × {ws.max_column} colonnes\n")

    # Extraire les données clés (format MT5 standard)
    print(f"{'='*70}")
    print(f"  DONNÉES BRUTES (30 premières lignes)")
    print(f"{'='*70}\n")

    for row_idx in range(1, min(31, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(11, ws.max_column + 1)):  # 10 premières colonnes
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_data.append(str(cell_value)[:30])  # Tronquer à 30 chars

        if row_data:  # Afficher seulement lignes non-vides
            print(f"Ligne {row_idx:2d}: {' | '.join(row_data)}")

    # Extraire métriques (format français MT5)
    print(f"\n{'='*70}")
    print(f"  MÉTRIQUES BACKTEST")
    print(f"{'='*70}\n")

    metrics = {}

    # Scanner toutes les lignes à la recherche de patterns français MT5
    for row_idx in range(1, ws.max_row + 1):
        row_text = []
        for col_idx in range(1, min(6, ws.max_column + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                row_text.append(str(val))

        line = " ".join(row_text).lower()

        # Profit net total
        if "profit net total" in line or "total net profit" in line:
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, (int, float)) and val != 0:
                    metrics["Net Profit"] = val
                    print(f"✓ Profit Net: ${val:.2f}")
                    break

        # Bénéfice brut / Gross Profit
        if "bénéfice brut" in line or "profit brut" in line or "gross profit" in line:
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, (int, float)) and val > 0:
                    metrics["Gross Profit"] = val
                    print(f"✓ Bénéfice Brut: ${val:.2f}")
                    break

        # Perte brute / Gross Loss
        if "perte brute" in line or "total loss" in line or "gross loss" in line:
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, (int, float)) and val < 0:
                    metrics["Gross Loss"] = val
                    print(f"✓ Perte Brute: ${val:.2f}")
                    break

        # Facteur de profit
        if "facteur de profit" in line or "profit factor" in line:
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, (int, float)) and 0 < val < 100:
                    metrics["Profit Factor"] = val
                    print(f"✓ Profit Factor: {val:.2f}")
                    break

        # Facteur de récupération
        if "facteur de récupération" in line or "recovery factor" in line:
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, (int, float)):
                    metrics["Recovery Factor"] = val
                    print(f"✓ Recovery Factor: {val:.2f}")
                    break

        # Total trades
        if "nombre total d'opérations" in line or "total deals" in line or "total trades" in line:
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, int) and val > 0:
                    metrics["Total Trades"] = val
                    print(f"✓ Nombre Total Trades: {val}")
                    break

        # Trades gagnants
        if "transactions rentables" in line or "profitable trades" in line or "opérations rentables" in line:
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, str) and "(" in val:
                    # Format: "12 (44.44%)"
                    win_trades = int(val.split("(")[0].strip())
                    win_pct = float(val.split("(")[1].replace("%", "").replace(")", "").strip())
                    metrics["Win Trades"] = win_trades
                    metrics["Win Rate"] = win_pct
                    print(f"✓ Trades Gagnants: {win_trades} ({win_pct:.1f}%)")
                    break

        # Max drawdown
        if "prélèvement maximal" in line or "max drawdown" in line or "maximal drawdown" in line:
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, str) and "(" in val:
                    # Format: "41.41 (8.28%)"
                    dd_abs = float(val.split("(")[0].strip())
                    dd_pct = float(val.split("(")[1].replace("%", "").replace(")", "").strip())
                    metrics["Max DD Abs"] = dd_abs
                    metrics["Max DD %"] = dd_pct
                    print(f"✓ Max Drawdown: ${dd_abs:.2f} ({dd_pct:.2f}%)")
                    break

        # Balance finale
        if "solde final" in line or "balance final" in line or "balance end" in line:
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, (int, float)) and val > 100:
                    metrics["Final Balance"] = val
                    print(f"✓ Balance Finale: ${val:.2f}")
                    break

    if not metrics:
        print("⚠️ Aucune métrique détectée. Affichage des 80 premières lignes...")
        for row_idx in range(1, min(81, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(8, ws.max_column + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    row_data.append(str(val)[:40])
            if row_data:
                print(f"L{row_idx:3d}: {' | '.join(row_data)}")

    # Chercher tableau de trades
    print(f"\n{'='*70}")
    print(f"  RECHERCHE TABLEAU DES TRADES")
    print(f"{'='*70}\n")

    trade_headers = ["ticket", "time", "type", "volume", "price", "s/l", "t/p", "profit"]

    for row_idx in range(1, min(50, ws.max_row + 1)):
        row_text = " ".join([
            str(ws.cell(row=row_idx, column=col_idx).value or "").lower()
            for col_idx in range(1, ws.max_column + 1)
        ])

        if any(header in row_text for header in trade_headers[:3]):
            print(f"✓ Tableau de trades trouvé à la ligne {row_idx}")
            print(f"  Headers: {row_text[:100]}")

            # Afficher quelques trades
            print(f"\n  Premiers trades:")
            for i in range(1, 11):
                trade_row = row_idx + i
                if trade_row <= ws.max_row:
                    trade_data = [
                        str(ws.cell(row=trade_row, column=col_idx).value or "")
                        for col_idx in range(1, min(9, ws.max_column + 1))
                    ]
                    if any(trade_data):
                        print(f"    Trade {i}: {' | '.join(trade_data)}")
            break
    else:
        print("⚠️ Tableau de trades non trouvé")

    print(f"\n{'='*70}")

    wb.close()

    return metrics


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_backtest_excel.py <fichier.xlsx>")
        sys.exit(1)

    excel_file = sys.argv[1]

    if not Path(excel_file).exists():
        print(f"❌ Fichier introuvable: {excel_file}")
        sys.exit(1)

    try:
        metrics = extract_backtest_summary(excel_file)
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Erreur: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
